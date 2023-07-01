import requests
from bs4 import BeautifulSoup as bs
import openpyxl
from openpyxl import load_workbook
import re
import firebase_admin

from firebase_admin import credentials
from firebase_admin import db

cred = credentials.Certificate(
    "D:\Projects\pythonProjects\SUMTableExcel\dule-3fb37-firebase-adminsdk-fc53p-ea7d2ae02a.json")
firebase_admin.initialize_app(cred, {

    'databaseURL': 'https://dule-3fb37-default-rtdb.europe-west1.firebasedatabase.app/'
})


def clean_text(text):
    cleaned_text = re.sub('[^а-яА-Я0-9\s]+', '', text)
    for i, c in enumerate(cleaned_text):
        if c.isupper():
            cleaned_text = cleaned_text[:i]
    if len(cleaned_text) < 1:
        return text[40:]
    return cleaned_text


def contains_numbers(string):
    # определяем регулярное выражение, которое будет искать числа
    pattern = r'\d'
    # ищем совпадения в строке с помощью метода search модуля re
    match = re.search(pattern, string)
    # возвращаем True, если было найдено совпадение, и False, если нет
    return bool(match)


url = 'https://guu.ru/student/schedule/'
user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
headers = {'User-Agent': user_agent}
response = requests.get(url, headers=headers)
soup = bs(response.content, "html.parser")


lst_files = []
# ref.set('/')
files = soup.find_all('a', attrs={'class':'doc-unit odd'})
text_to_match = ['аспир', 'сессия']
for i, file in enumerate(files):
    if any(text in clean_text(file["href"]) for text in text_to_match):
        continue
    print(file['href'])
    reqfile = requests.get(file['href'],  headers=headers)
    if 'xlsx' in file['href']:
        with open (f'{clean_text(file["href"])}.xlsx', 'wb') as f:
            f.write(reqfile.content)
            print(f'{clean_text(file["href"])}.xlsx')
            workbook = openpyxl.load_workbook(f'{clean_text(file["href"])}.xlsx')
            for sheet in workbook.worksheets:
                for merged_cell in sheet.merged_cells.__copy__():
                    value = sheet.cell(row=merged_cell.min_row, column=merged_cell.min_col).value
                    min_row, min_col, max_row, max_col = merged_cell.min_row, merged_cell.min_col, merged_cell.max_row, merged_cell.max_col
                    sheet.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
                    for row in range(merged_cell.min_row, merged_cell.max_row + 1):
                        for column in range(merged_cell.min_col, merged_cell.max_col + 1):
                            sheet.cell(row=row, column=column).value = value
                workbook.save(f'{clean_text(file["href"])}.xlsx')
                workbook.close()
        lst_files.append(f'{clean_text(file["href"])}.xlsx')
data = {}
ref = db.reference('/')
ref.set('')

print(lst_files)
for file in lst_files:
    file_name = file.replace('.xlsx', '').replace('//', '')

    workbook = load_workbook(filename=file)
    for i, worksheet in enumerate(workbook.worksheets):
        sheet_name = workbook.sheetnames[i]
        print(sheet_name)
        if sheet_name != '3 курс':
            for g in range(150):
                item = worksheet.cell(6, 6 + g).value
                if (str(item).__len__() > 6):
                    name_group = f'{worksheet.cell(6, 6 + g).value} {worksheet.cell(7, 6 + g).value}'
                    name_group = name_group.replace('\n', '')
                    time_column = 3
                    for c_day in range(1, 10):
                        if str(worksheet.cell(8, c_day).value).lower() == 'день':
                            time_column = c_day + 1

                    for p in range(48):
                        lesson = f'{worksheet.cell(9 + p, 6 + g).value}'
                        time = f'{worksheet.cell(9 + p, time_column).value}'
                        day = f'{worksheet.cell(9 + p, time_column - 1).value}'
                        week = f'{worksheet.cell(9 + p, time_column + 1).value}'

                        if sheet_name not in data:
                            data[sheet_name] = {}
                        if name_group not in data[sheet_name]:
                            data[sheet_name][name_group] = {}
                        data[sheet_name][name_group][f'lesson_{p}'] = {
                            'text': lesson,
                            'time': time,
                            'day': day,
                            'week': week
                        }
                        print(sheet_name,name_group,p)
                        ref.update(data)
                        # ref.child(sheet_name).child(name_group).child(f'lesson_{p}').child(f'text').set(lesson)
                        # ref.child(sheet_name).child(name_group).child(f'lesson_{p}').child(f'time').set(time)
                        # ref.child(sheet_name).child(name_group).child(f'lesson_{p}').child(f'day').set(day)
                        # ref.child(sheet_name).child(name_group).child(f'lesson_{p}').child(f'week').set(week)
# на магистратуре трабл
workbook.close()


